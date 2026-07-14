import csv
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from typing import Any
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree


class SpreadsheetReadError(ValueError):
    pass


@dataclass
class SpreadsheetData:
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclass(frozen=True)
class MergeRange:
    start_row: int
    start_col: int
    end_row: int
    end_col: int


@dataclass
class WorksheetRow:
    row_number: int
    values: list[Any]
    merged_ranges: dict[int, MergeRange]

    def __iter__(self):
        yield self.row_number
        yield self.values


NS_MAIN = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def read_spreadsheet(uploaded_file, sheet_name: str = "") -> SpreadsheetData:
    filename = uploaded_file.name.lower()
    content = uploaded_file.read()
    if filename.endswith(".csv"):
        table = _read_csv(content)
        return _table_to_data(table, sheet_name="CSV")
    if filename.endswith(".xlsx"):
        return _read_xlsx(content, sheet_name=sheet_name.strip())
    raise SpreadsheetReadError("Formato não suportado.")


def read_exercise_tracking_spreadsheet(uploaded_file, sheet_name: str = "") -> SpreadsheetData:
    original_filename = uploaded_file.name
    filename = original_filename.lower()
    content = uploaded_file.read()
    if not filename.endswith(".xlsx"):
        raise SpreadsheetReadError("O historico de exercicios deve ser importado em XLSX.")

    try:
        with ZipFile(BytesIO(content)) as archive:
            sheets = _workbook_sheets(archive)
            if not sheets:
                raise SpreadsheetReadError("Nenhuma aba encontrada no XLSX.")

            selected_sheets = [_select_sheet(sheets, sheet_name.strip())] if sheet_name.strip() else sheets
            shared_strings = _shared_strings(archive)
            rows: list[dict[str, Any]] = []
            parsed_sheet_names = []

            for selected in selected_sheets:
                table = _worksheet_numbered_rows(archive, selected["path"], shared_strings)
                sheet_rows = _exercise_tracking_rows(table, selected["name"])
                if sheet_rows:
                    parsed_sheet_names.append(selected["name"])
                    rows.extend(sheet_rows)

            if not rows:
                raise SpreadsheetReadError("Nenhum exercicio foi encontrado nas abas selecionadas.")

            _harmonize_tracking_patient_names(rows, _tracking_filename_stem(original_filename))

            return SpreadsheetData(
                sheet_name=parsed_sheet_names[0] if len(parsed_sheet_names) == 1 else "Todas as abas",
                headers=[
                    "paciente_nome",
                    "categoria",
                    "exercicio",
                    "marcacoes",
                    "sheet_name",
                    "linha_origem",
                ],
                rows=rows,
            )
    except BadZipFile as exc:
        raise SpreadsheetReadError("Arquivo XLSX inválido.") from exc


def _read_csv(content: bytes) -> list[list[Any]]:
    text = _decode_csv(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(StringIO(text), dialect)]


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SpreadsheetReadError("Não foi possível ler o CSV.")


def _read_xlsx(content: bytes, sheet_name: str = "") -> SpreadsheetData:
    openpyxl_data = _read_xlsx_with_openpyxl(content, sheet_name=sheet_name)
    if openpyxl_data is not None:
        return openpyxl_data
    return _read_xlsx_with_stdlib(content, sheet_name=sheet_name)


def _read_xlsx_with_openpyxl(content: bytes, sheet_name: str = "") -> SpreadsheetData | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SpreadsheetReadError(f"Não foi possível ler o XLSX: {exc}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise SpreadsheetReadError(f"Aba '{sheet_name}' não encontrada.")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    table = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return _table_to_data(table, sheet_name=worksheet.title)


def _read_xlsx_with_stdlib(content: bytes, sheet_name: str = "") -> SpreadsheetData:
    try:
        with ZipFile(BytesIO(content)) as archive:
            sheets = _workbook_sheets(archive)
            if not sheets:
                raise SpreadsheetReadError("Nenhuma aba encontrada no XLSX.")

            selected = _select_sheet(sheets, sheet_name)
            shared_strings = _shared_strings(archive)
            table = _worksheet_rows(archive, selected["path"], shared_strings)
            return _table_to_data(table, sheet_name=selected["name"])
    except BadZipFile as exc:
        raise SpreadsheetReadError("Arquivo XLSX inválido.") from exc


def _workbook_sheets(archive: ZipFile) -> list[dict[str, str]]:
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationships = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall("r:Relationship", NS_REL)
        if rel.attrib.get("Type", "").endswith("/worksheet")
    }

    sheets = []
    for sheet in workbook_root.findall("x:sheets/x:sheet", NS_MAIN):
        rel_id = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id")
        target = relationships.get(rel_id)
        if not target:
            continue
        sheets.append(
            {
                "name": sheet.attrib["name"],
                "path": _xlsx_target_path(target),
            }
        )
    return sheets


def _select_sheet(sheets: list[dict[str, str]], sheet_name: str) -> dict[str, str]:
    if not sheet_name:
        return sheets[0]
    for sheet in sheets:
        if sheet["name"].casefold() == sheet_name.casefold():
            return sheet
    raise SpreadsheetReadError(f"Aba '{sheet_name}' não encontrada.")


def _xlsx_target_path(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return str(PurePosixPath("xl") / target)


def _shared_strings(archive: ZipFile) -> list[str]:
    try:
        root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    values = []
    for item in root.findall("x:si", NS_MAIN):
        values.append("".join(item.itertext()))
    return values


def _worksheet_rows(archive: ZipFile, path: str, shared_strings: list[str]) -> list[list[Any]]:
    return [values for _row_number, values in _worksheet_numbered_rows(archive, path, shared_strings)]


def _worksheet_numbered_rows(
    archive: ZipFile,
    path: str,
    shared_strings: list[str],
) -> list[WorksheetRow]:
    root = ElementTree.fromstring(archive.read(path))
    cell_values: dict[tuple[int, int], Any] = {}
    row_numbers: set[int] = set()

    for row in root.findall("x:sheetData/x:row", NS_MAIN):
        row_number = int(row.attrib.get("r", len(row_numbers) + 1))
        row_numbers.add(row_number)
        next_column = 1
        for cell in row.findall("x:c", NS_MAIN):
            column = _cell_column_index(cell.attrib.get("r", "")) or next_column
            cell_values[(row_number, column)] = _cell_value(cell, shared_strings)
            next_column = column + 1

    virtual_ranges: dict[tuple[int, int], MergeRange] = {}
    for merge_range in _worksheet_merged_ranges(root):
        top_left_value = cell_values.get((merge_range.start_row, merge_range.start_col), "")
        if not _present(top_left_value):
            continue
        for row_number in range(merge_range.start_row, merge_range.end_row + 1):
            row_numbers.add(row_number)
            for column in range(merge_range.start_col, merge_range.end_col + 1):
                key = (row_number, column)
                if key == (merge_range.start_row, merge_range.start_col):
                    continue
                if not _present(cell_values.get(key, "")):
                    cell_values[key] = top_left_value
                    virtual_ranges[key] = merge_range

    rows: list[WorksheetRow] = []
    for row_number in sorted(row_numbers):
        columns = [column for current_row, column in cell_values if current_row == row_number]
        max_column = max(columns) if columns else 0
        values = [cell_values.get((row_number, column), "") for column in range(1, max_column + 1)]
        merged_ranges = {
            column - 1: virtual_ranges[(row_number, column)]
            for column in range(1, max_column + 1)
            if (row_number, column) in virtual_ranges
        }
        rows.append(WorksheetRow(row_number=row_number, values=values, merged_ranges=merged_ranges))
    return rows



def _worksheet_merged_ranges(root) -> list[MergeRange]:
    ranges = []
    for merge_cell in root.findall("x:mergeCells/x:mergeCell", NS_MAIN):
        reference = merge_cell.attrib.get("ref", "")
        merge_range = _cell_range(reference)
        if merge_range is not None:
            ranges.append(merge_range)
    return ranges


def _cell_range(reference: str) -> MergeRange | None:
    if not reference:
        return None
    parts = reference.split(":", 1)
    start = _cell_position(parts[0])
    end = _cell_position(parts[1] if len(parts) > 1 else parts[0])
    if start is None or end is None:
        return None
    start_row, start_col = start
    end_row, end_col = end
    return MergeRange(
        start_row=min(start_row, end_row),
        start_col=min(start_col, end_col),
        end_row=max(start_row, end_row),
        end_col=max(start_col, end_col),
    )


def _cell_position(reference: str) -> tuple[int, int] | None:
    match = re.match(r"([A-Z]+)(\d+)", reference.upper())
    if not match:
        return None
    return int(match.group(2)), _column_letters_to_index(match.group(1))


def _column_letters_to_index(letters: str) -> int:
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def _cell_column_index(reference: str) -> int | None:
    match = re.match(r"([A-Z]+)", reference.upper())
    if not match:
        return None
    return _column_letters_to_index(match.group(1))


def _cell_value(cell, shared_strings: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        inline = cell.find("x:is", NS_MAIN)
        return "" if inline is None else "".join(inline.itertext())

    value = cell.find("x:v", NS_MAIN)
    if value is None or value.text is None:
        return ""

    raw_value = value.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (ValueError, IndexError):
            return raw_value
    if cell_type == "b":
        return raw_value == "1"
    if cell_type in {"str", "e"}:
        return raw_value
    return _coerce_number(raw_value)


def _coerce_number(value: str) -> Any:
    try:
        number = float(value)
    except ValueError:
        return value
    if number.is_integer():
        return int(number)
    return number


def _table_to_data(table: list[list[Any]], sheet_name: str) -> SpreadsheetData:
    normalized_rows = [row for row in table if any(_present(cell) for cell in row)]
    if not normalized_rows:
        raise SpreadsheetReadError("A planilha está vazia.")

    headers = [_string_header(value, index) for index, value in enumerate(normalized_rows[0], start=1)]
    headers = _unique_headers(headers)
    rows = []
    for row in normalized_rows[1:]:
        values = {header: row[index] if index < len(row) else "" for index, header in enumerate(headers)}
        if any(_present(value) for value in values.values()):
            rows.append(values)

    if not rows:
        raise SpreadsheetReadError("A planilha não possui linhas de dados.")
    return SpreadsheetData(sheet_name=sheet_name, headers=headers, rows=rows)


def _string_header(value: Any, index: int) -> str:
    text = "" if value is None else str(value).strip()
    return text or f"coluna_{index}"


def _unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique = []
    for header in headers:
        count = seen.get(header, 0)
        seen[header] = count + 1
        unique.append(header if count == 0 else f"{header}_{count + 1}")
    return unique


def _present(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _exercise_tracking_rows(numbered_rows: list[WorksheetRow], sheet_name: str) -> list[dict[str, Any]]:
    date_row_position, date_columns = _find_tracking_date_columns(numbered_rows)
    if date_row_position is None:
        return []

    rows_before_header = numbered_rows[:date_row_position]
    patient_name = _tracking_patient_name(rows_before_header)
    if not patient_name:
        raise SpreadsheetReadError(f"Aba '{sheet_name}' sem nome de paciente antes das datas.")

    output_rows: list[dict[str, Any]] = []
    current_category = ""

    for worksheet_row in numbered_rows[date_row_position + 1 :]:
        row_number = worksheet_row.row_number
        row = worksheet_row.values
        category_text = _tracking_text(_row_value(row, 1))
        exercise_name = _tracking_text(_row_value(row, 2))
        if _tracking_merged_category_header(worksheet_row, category_text, exercise_name):
            exercise_name = ""

        if _is_tracking_observation_label(category_text) or _is_tracking_observation_label(exercise_name):
            break

        if category_text:
            current_category = category_text

        if not exercise_name:
            continue

        category_name = category_text or current_category
        if not category_name:
            raise SpreadsheetReadError(
                f"Aba '{sheet_name}', linha {row_number}: categoria nao encontrada para exercicio '{exercise_name}'."
            )

        marks = []
        for column_index, performed_date in date_columns:
            if _worksheet_virtual_merge(worksheet_row, column_index):
                continue
            mark = _tracking_text(_row_value(row, column_index))
            if not mark:
                continue
            marks.append({"data": performed_date, "marca": mark})

        output_rows.append(
            {
                "paciente_nome": patient_name,
                "categoria": category_name,
                "exercicio": exercise_name,
                "marcacoes": marks,
                "sheet_name": sheet_name,
                "linha_origem": row_number,
            }
        )

    return output_rows



def _tracking_merged_category_header(worksheet_row: WorksheetRow, category_text: str, exercise_name: str) -> bool:
    category_column = 2
    exercise_column = 3
    merge_range = worksheet_row.merged_ranges.get(exercise_column - 1)
    if _tracking_horizontal_header_merge(merge_range, category_column, exercise_column):
        return True

    next_cell_merge = worksheet_row.merged_ranges.get(exercise_column)
    return (
        _tracking_header_text_matches_category(category_text, exercise_name)
        and _tracking_horizontal_header_merge(next_cell_merge, exercise_column, exercise_column + 1)
    )


def _tracking_horizontal_header_merge(
    merge_range: MergeRange | None,
    start_column_limit: int,
    covered_column: int,
) -> bool:
    if merge_range is None:
        return False
    return (
        merge_range.start_row == merge_range.end_row
        and merge_range.start_col <= start_column_limit
        and merge_range.end_col >= covered_column
    )


def _tracking_header_text_matches_category(category_text: str, exercise_name: str) -> bool:
    category_key = _tracking_normalized_text(category_text)
    exercise_key = _tracking_normalized_text(exercise_name)
    if not category_key or not exercise_key:
        return False
    return exercise_key == category_key or exercise_key.startswith(f"{category_key}_")


def _worksheet_virtual_merge(worksheet_row: WorksheetRow, column_index: int) -> bool:
    return column_index in worksheet_row.merged_ranges


def _find_tracking_date_columns(
    numbered_rows: list[WorksheetRow],
) -> tuple[int | None, list[tuple[int, date]]]:
    for row_position, worksheet_row in enumerate(numbered_rows):
        row = worksheet_row.values
        date_columns = [
            (column_index, parsed_date)
            for column_index, value in enumerate(row)
            if not _worksheet_virtual_merge(worksheet_row, column_index)
            if (parsed_date := _parse_tracking_header_date(value)) is not None
        ]
        if date_columns and any(column_index >= 3 for column_index, _parsed_date in date_columns):
            return row_position, date_columns
    return None, []


def _parse_tracking_header_date(value: Any) -> date | None:
    if not _present(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 30000 <= float(value) <= 80000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    text = _tracking_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _tracking_patient_name(numbered_rows: list[tuple[int, list[Any]]]) -> str:
    for _row_number, row in numbered_rows:
        for column_index in (1, 0, 2):
            text = _tracking_text(_row_value(row, column_index))
            if _is_tracking_patient_name(text):
                return text
    return ""


def _harmonize_tracking_patient_names(rows: list[dict[str, Any]], filename_stem: str) -> None:
    patient_names = sorted({_tracking_text(row.get("paciente_nome")) for row in rows if row.get("paciente_nome")})
    if not patient_names:
        return

    filename_candidate = _tracking_text(filename_stem)
    if (
        filename_candidate
        and len(filename_candidate) > max(len(name) for name in patient_names)
        and _tracking_name_contains_all(filename_candidate, patient_names)
    ):
        canonical_name = filename_candidate
    elif len(patient_names) < 2:
        return
    else:
        canonical_name = max(patient_names, key=len)
        if not _tracking_name_contains_all(canonical_name, patient_names):
            return

    for row in rows:
        row["paciente_nome"] = canonical_name


def _tracking_name_contains_all(candidate: str, patient_names: list[str]) -> bool:
    normalized_candidate = _tracking_normalized_text(candidate)
    return all(_tracking_normalized_text(name) in normalized_candidate for name in patient_names)


def _tracking_filename_stem(filename: str) -> str:
    stem = PurePosixPath(filename).name
    stem = re.sub(r"\.xlsx$", "", stem, flags=re.IGNORECASE)
    return re.sub(r"\(\d+\)$", "", stem).strip()


def _is_tracking_patient_name(text: str) -> bool:
    if not text or text == "-":
        return False
    if _is_tracking_observation_label(text):
        return False
    if len(text) > 80:
        return False
    if re.match(r"^\d{1,2}/\d{1,2}", text):
        return False
    if ":" in text:
        return False
    return True


def _is_tracking_observation_label(text: str) -> bool:
    normalized = _tracking_normalized_text(text)
    return normalized in {"observacao", "observacoes"}


def _tracking_normalized_text(text: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_text.strip().lower())
    return normalized.strip("_")


def _tracking_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _row_value(row: list[Any], column_index: int) -> Any:
    return row[column_index] if column_index < len(row) else ""
